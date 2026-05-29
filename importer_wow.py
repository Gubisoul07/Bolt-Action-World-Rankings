"""WoW importer — reads Excel files from the Data/ folder and upserts into
the rankings DB.

Supported editions:
  2024 — Data/WOW 2024 data.xlsx  (all rounds on a single 'All results' sheet)
  2025 — Data/WOW 2025 results.xlsx  (one 'Round N' sheet per round)

Called by app.py's /admin/import/wow/<year> route; can also be run standalone:
  python importer_wow.py 2024
  python importer_wow.py 2025
"""
import os
import re
import sys

import openpyxl

from app import get_db, get_cursor, insert_returning_id, q

DATA_DIR = os.path.join(os.path.dirname(__file__), 'Data')

# Column offsets for each group block (up to 4 groups side-by-side, 10 cols each).
_GROUP_OFFSETS = [0, 10, 20, 30]

_WORD_TO_NUM = {
    'one': 1, 'two': 2, 'three': 3, 'four': 4,
    'five': 5, 'six': 6, 'seven': 7,
}

# Cells in col[offset+0] that are NOT table identifiers.
_SKIP_CELLS = {'table', 'bonus', 'group a', 'group b', 'group c', 'group d',
               'ranking', 'none'}

# Values in the player-name columns that indicate a header row, not player data.
_SKIP_NAMES = {'army', 'points', 'secondary', 'roll off winner and choice', ''}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _detect_round(row):
    """Return round number if this row is a round header, else None."""
    for col_idx in (0, 1):
        if col_idx >= len(row):
            break
        cell = str(row[col_idx] or '').strip().lower()
        if not cell.startswith('round'):
            continue
        m = re.search(r'\d+', cell)
        if m:
            return int(m.group())
        for word, num in _WORD_TO_NUM.items():
            if word in cell:
                return num
    return None


def _pts_to_outcome(pts_a, pts_b):
    """Convert (points_a, points_b) → W/D/L from player_a's view, or None."""
    if pts_a is None and pts_b is None:
        return None
    try:
        pa = float(pts_a if pts_a is not None else 0)
        pb = float(pts_b if pts_b is not None else 0)
    except (TypeError, ValueError):
        return None
    if pa > pb:
        return 'W'
    if pa == pb:
        return 'D'
    return 'L'


def _is_player_row(row, offset):
    """True if this row contains individual player game data at the given offset."""
    if len(row) <= offset + 7:
        return False
    table_cell = row[offset]
    name_a     = row[offset + 1]
    name_b     = row[offset + 3]
    pts_a      = row[offset + 6]
    pts_b      = row[offset + 7]

    if str(table_cell or '').strip().lower() in _SKIP_CELLS:
        return False
    if not name_a or str(name_a).strip().lower() in _SKIP_NAMES:
        return False
    if not name_b or str(name_b).strip().lower() in _SKIP_NAMES:
        return False
    # At least one points cell must be a real number.
    if not isinstance(pts_a, (int, float)) and not isinstance(pts_b, (int, float)):
        return False
    return True


def _table_to_board(table_cell, counters, round_no):
    """Convert a table cell (letter, number, or None) to an integer board number.

    For rounds where no table label is recorded (some 2024 rounds), a
    synthetic counter per round is used starting at 101.
    """
    s = str(table_cell or '').strip()
    if s:
        try:
            return int(s)
        except ValueError:
            first = s[0].upper()
            if 'A' <= first <= 'Z':
                return ord(first) - ord('A') + 1
    # Synthetic fallback.
    counters.setdefault(round_no, 100)
    counters[round_no] += 1
    return counters[round_no]


def _extract_games(ws, fixed_round=None):
    """Yield (round_no, board_number, name_a, name_b, outcome) from a worksheet.

    fixed_round — pass an int when the round number is known from the sheet
                  name (2025 format).  Pass None to detect it from row headers
                  (2024 format).
    """
    games = []
    current_round = fixed_round or 0
    counters = {}  # synthetic board-number counters, keyed by round_no

    for row in ws.iter_rows(values_only=True):
        # Detect round-header rows only when round is not fixed.
        if fixed_round is None:
            detected = _detect_round(row)
            if detected:
                current_round = detected
                continue

        if current_round == 0:
            continue

        for offset in _GROUP_OFFSETS:
            if not _is_player_row(row, offset):
                continue
            name_a  = str(row[offset + 1]).strip()
            name_b  = str(row[offset + 3]).strip()
            outcome = _pts_to_outcome(row[offset + 6], row[offset + 7])
            if not outcome:
                continue
            board = _table_to_board(row[offset], counters, current_round)
            games.append((current_round, board, name_a, name_b, outcome))

    return games


def _load_all_games(year):
    """Parse the Excel file(s) for *year* and return a list of game tuples."""
    if year == 2024:
        path = os.path.join(DATA_DIR, 'WOW 2024 data.xlsx')
        wb   = openpyxl.load_workbook(path, data_only=True)
        games = _extract_games(wb['All results'], fixed_round=None)
        wb.close()

    elif year == 2025:
        path = os.path.join(DATA_DIR, 'WOW 2025 results.xlsx')
        wb   = openpyxl.load_workbook(path, data_only=True)
        games = []
        for round_no in range(1, 7):
            sheet_name = f'Round {round_no}'
            if sheet_name in wb.sheetnames:
                games.extend(_extract_games(wb[sheet_name], fixed_round=round_no))
        wb.close()

    else:
        raise ValueError(f'No WoW data configured for year {year}')

    return games


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def import_wow(year: int) -> dict:
    """Import WoW data for *year* from the Data/ folder.  Returns stats dict."""
    games = _load_all_games(year)

    conn = get_db()
    c    = get_cursor(conn)
    stats = {'editions': 0, 'players': 0, 'games': 0, 'new_competitors': 0}

    # Upsert source_edition (synthetic ID = year, since WoW has no authoritative ID).
    edition_name = f'WoW {year}'
    c.execute(q('SELECT id FROM source_edition WHERE source=? AND source_edition_id=?'),
              ('wow', year))
    row = c.fetchone()
    if row:
        sed_id = row['id']
        c.execute(q('UPDATE source_edition SET year=?, name=?, match_data_available=1 WHERE id=?'),
                  (year, edition_name, sed_id))
    else:
        sed_id = insert_returning_id(c, q(
            '''INSERT INTO source_edition
               (source, source_edition_id, year, name, match_data_available)
               VALUES (?, ?, ?, ?, 1)'''),
            ('wow', year, year, edition_name))
        stats['editions'] += 1

    # Build source_competitor map: name → competitor_id.
    all_names = set()
    for _, _, name_a, name_b, _ in games:
        all_names.add(name_a)
        all_names.add(name_b)

    name_to_comp = {}
    for name in sorted(all_names):
        source_key = name.lower()
        c.execute(q('SELECT id, competitor_id FROM source_competitor WHERE source=? AND source_key=?'),
                  ('wow', source_key))
        sc_row = c.fetchone()
        if sc_row:
            name_to_comp[name] = sc_row['competitor_id']
            # Keep display name fresh.
            c.execute(q('UPDATE source_competitor SET name=? WHERE id=?'),
                      (name, sc_row['id']))
        else:
            comp_id = insert_returning_id(
                c, q('INSERT INTO competitor (display_name) VALUES (?)'), (name,))
            insert_returning_id(c, q(
                '''INSERT INTO source_competitor
                   (source, source_key, name, competitor_id)
                   VALUES (?, ?, ?, ?)'''),
                ('wow', source_key, name, comp_id))
            name_to_comp[name] = comp_id
            stats['new_competitors'] += 1
            stats['players'] += 1

    # Upsert games.
    for round_no, board_number, name_a, name_b, outcome in games:
        ca = name_to_comp.get(name_a)
        cb = name_to_comp.get(name_b)
        if ca is None or cb is None:
            continue

        c.execute(q('''SELECT id FROM game
                       WHERE source_edition_id=? AND round_order=?
                       AND board_number=? AND competitor_a=? AND competitor_b=?'''),
                  (sed_id, round_no, board_number, ca, cb))
        if c.fetchone():
            c.execute(q('''UPDATE game SET outcome=?
                           WHERE source_edition_id=? AND round_order=?
                           AND board_number=? AND competitor_a=? AND competitor_b=?'''),
                      (outcome, sed_id, round_no, board_number, ca, cb))
        else:
            insert_returning_id(c, q(
                '''INSERT INTO game
                   (source_edition_id, round_order, board_number,
                    competitor_a, competitor_b, outcome, dice_kills_a, dice_kills_b)
                   VALUES (?, ?, ?, ?, ?, ?, 0, 0)'''),
                (sed_id, round_no, board_number, ca, cb, outcome))
            stats['games'] += 1

    conn.commit()
    conn.close()
    return stats


if __name__ == '__main__':
    year = int(sys.argv[1]) if len(sys.argv) > 1 else 2024
    print(f'Importing WoW {year} …')
    result = import_wow(year)
    print(result)
